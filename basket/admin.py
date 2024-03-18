import logging
import re

from django.conf import settings
from django.contrib import admin, messages
from openpyxl import load_workbook, Workbook
from django_mail_admin.models import OutgoingEmail, EmailTemplate, PRIORITY
from django_mail_admin import mail

from basket import models
from basket.forms import MotorcycleAdminForm, SendInfoEmailForm
from basket.models import SendEmailAtCheckoutCartModel


@admin.register(models.Category)
class AdminCategory(admin.ModelAdmin):
    list_display = ['name']


class EngineInline(admin.TabularInline):
    model = models.Engine
    extra = 1
    classes = ['collapse']


@admin.register(models.Motorcycle)
class AdminMotorcycle(admin.ModelAdmin):
    list_display = ['name', 'category']
    prepopulated_fields = {'slug': ['name']}
    inlines = (EngineInline,)

    def get_form(self, request, obj=None, change=False, **kwargs):
        kwargs['form'] = MotorcycleAdminForm
        return super().get_form(request, obj, **kwargs)
    
    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        mainparts = []

        if obj.excel_file:
            try:
                wb: Workbook = load_workbook(obj.excel_file)
                ws = wb[wb.sheetnames[0]]

                counter_mainparts = -1
                previous_value = 0

                # Получаем данные с Excel файла и создаем объект экземпляра MainPart
                for row in ws.iter_rows(min_row=1, min_col=2, max_col=7):
                    value_mainpart_or_number = str(row[0].value)
                    try:
                        value_number = value_mainpart_or_number
                        code_value = str(row[1].value)
                        name_value = str(row[3].value)
                        quantity = int(row[4].value)

                        if code_value.startswith('R'):
                            part = models.Part.objects.create(
                                main_part=mainparts[counter_mainparts],
                                number=value_number,
                                code=code_value,
                                name=name_value,
                            )

                            if row[5].value is not None:
                                models.NotePartModel.objects.create(
                                    part=part,
                                    note=str(row[5].value)
                                )

                            if quantity > 1:
                                models.NotePartModel.objects.create(
                                    part=part,
                                    note=f'Для заказа комплекта необходимо {quantity} шт'
                                )

                    except:
                        mainpart_value = str(value_mainpart_or_number)
                        if mainpart_value != 'None' and any(mainpart_value.startswith(str(i) + '.') for i in range(1, 101)):
                            # Извлечение номера с помощью regex
                            match = re.match(r'^(\d+)\.', mainpart_value)

                            if match:
                                number_mainpart = int(match.group(1))
                                mainpart_obj, created = models.MainPart.objects.get_or_create(
                                    motorcycle=obj,
                                    name=mainpart_value,
                                    ordering=number_mainpart,
                                )
                                counter_mainparts += 1
                                mainparts.append(mainpart_obj)
            except Exception as ex:
                obj.delete()
                messages.error(request, "Не удалось получить данные с excel файла")

            if mainparts:
                imgs_for_mainparts_list = request.FILES.getlist('imgs_for_mainparts')
                for i in range(0, len(mainparts)):
                    mainparts[i].image = imgs_for_mainparts_list[i]
                    mainparts[i].save()


@admin.register(models.MainPart)
class AdminMainPart(admin.ModelAdmin):
    list_display = ['name', 'motorcycle', 'ordering']


@admin.register(models.Part)
class AdminPart(admin.ModelAdmin):
    list_display = ['name', 'code', 'main_part', 'price', 'number']


class OrderedPartInline(admin.TabularInline):
    model = models.OrderedPart
    extra = 1
    classes = ['collapse']


class AdditionalItemInline(admin.TabularInline):
    model = models.AdditionalItemToCheckoutCart
    extra = 1
    classes = ['collapse']


class CommentAdministratorForCheckoutCartInline(admin.TabularInline):
    model = models.CommentAdministratorForCheckoutCart
    extra = 1
    classes = ['collapse']


class DeliveryMethodInline(admin.TabularInline):
    model = models.DeliveryMethod
    classes = ['collapse']
    fields = ['name', 'price']


class SendEmailAtCheckoutCartModelInline(admin.TabularInline):
    model = models.SendEmailAtCheckoutCartModel
    classes = ['collapse']
    fields = ['outgoing_email', 'template']

    def get_readonly_fields(self, request, obj=None):
        if hasattr(obj, 'send_email_cart_order'):
            send_email_at_checkout_cart_obj = obj.send_email_cart_order
            if send_email_at_checkout_cart_obj.outgoing_email is None:

                # Генерируем html код для отображения запчастей в виде таблицы
                html_list_parts = '<table border="1">\n'
                html_list_parts += '<tr><th>Наименование</th><th>Цена</th><th>Кол-во</th></tr>\n'
                for ordered_part in obj.ordered_parts.all():
                    html_list_parts += f'<tr><td>{ordered_part.part.name}</td><td>{ordered_part.part.price} р.</td><td>{ordered_part.quantity}</td></tr>\n'

                if hasattr(obj, 'additional_item'):
                    for item in obj.additional_item.all():
                        html_list_parts += f'<tr><td>{item.name}</td><td>{item.price} р.</td><td>1</td></tr>\n'
                html_list_parts += '</table>'

                # Переменные, которые будут использоваться в шаблоне
                variable_dict = {
                    'order_number': obj.pk,
                    'order_total_price': obj.total_price,
                    'order_products': html_list_parts,
                }

                if hasattr(obj, 'delivery_method'):
                    variable_dict['total_price_with_delivery'] = obj.delivery_method.total_price_with_delivery
                    variable_dict['price_delivery'] = obj.delivery_method.price

                # Отправляем письмо
                outgoing_email = mail.send(
                    'ilyasablin000@yandex.ru',
                    obj.user.email,
                    template=obj.send_email_cart_order.template,
                    priority=PRIORITY.now,
                    variable_dict=variable_dict,
                    )

                # Обновляем и сохраняем объект SendEmailAtCheckoutCartModel
                send_email_at_checkout_cart_obj.outgoing_email = outgoing_email
                send_email_at_checkout_cart_obj.from_email = outgoing_email.from_email
                send_email_at_checkout_cart_obj.to = outgoing_email.to
                send_email_at_checkout_cart_obj.save()

            return ['outgoing_email', 'template']
        return []


@admin.register(models.CheckoutCart)
class CheckoutCartAdmin(admin.ModelAdmin):
    list_display = ('id_as_order_number', 'client', 'total_price', 'created_at', 'order_status')
    list_display_links = ['id_as_order_number', 'client']
    list_filter = ['order_status', 'created_at']
    search_fields = ['id']
    date_hierarchy = 'created_at'
    inlines = (SendEmailAtCheckoutCartModelInline, CommentAdministratorForCheckoutCartInline, DeliveryMethodInline, AdditionalItemInline, OrderedPartInline)
    change_form_template = 'admin/basket/CheckoutCart/change_form_custom.html'

    def changeform_view(self, request, object_id=None, form_url='', extra_context=None):
        response = super().changeform_view(request, object_id, form_url, extra_context)
        obj = self.get_object(request, object_id)

        if request.method == 'POST':
            price_from_obj = 0

            for order_part in obj.ordered_parts.all():
                price_part = order_part.part.price
                quantity_part = order_part.quantity
                price_with_quantity_and_part = price_part * quantity_part
                price_from_obj += price_with_quantity_and_part

            for item in obj.additional_item.all():
                if item:
                    price_from_obj += item.price

            if obj.total_price != price_from_obj:
                obj.total_price = price_from_obj
                obj.save()
        elif request.method == 'GET' and hasattr(obj, 'delivery_method'):
            price_with_delivery_method = obj.total_price + obj.delivery_method.price
            if obj.delivery_method.total_price_with_delivery != price_with_delivery_method:
                obj.delivery_method.total_price_with_delivery = price_with_delivery_method
                obj.delivery_method.save()
        return response


@admin.register(models.ExcelFileCatalog)
class AdminExcelFileCatalog(admin.ModelAdmin):
        def save_model(self, request, obj, form, change):
            super().save_model(request, obj, form, change)
            parts = []

            if obj.excel_file:
                try:
                    wb: Workbook = load_workbook(obj.excel_file)
                    ws = wb[wb.sheetnames[0]]

                    for row in ws.iter_rows(min_row=1, min_col=1, max_col=3):
                        if row[0].value is not None and str(row[0].value).startswith('R'):
                            code_value = str(row[0].value)
                            price_value = str(row[2].value)
                            try:
                                part = models.Part.objects.filter(
                                    code=code_value
                                ).update(price=price_value)
                            except models.Part.DoesNotExist as ex:
                                logging.error(f'An error occurred while updating the price: {ex}')
                except FileNotFoundError as ex:
                    logging.error(f'Произошла ошибка при загрузке файла Excel: {ex}')


@admin.register(models.Engine)
class AdminEngine(admin.ModelAdmin):
    list_display = ['url']

"""
{{ order_number }}
{{ order_products }}
{{ order_total_price }}
{{ total_price_with_delivery }}
{{ price_delivery }}
"""

